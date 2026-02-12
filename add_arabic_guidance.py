#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script to add comprehensive Arabic guidance rows to Excel template files.
This script inserts a new row (row 2) with Arabic instructions for each column
based on intelligent column name detection.

Usage:
    python3 add_arabic_guidance.py

Requirements:
    pip install openpyxl
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import re
import glob


def contains_arabic(text):
    """Check if text contains Arabic characters."""
    if text is None:
        return False
    arabic_pattern = re.compile(r'[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF]+')
    return bool(arabic_pattern.search(str(text)))


def has_arabic_guidance_row(ws):
    """
    Check if row 2 already contains Arabic guidance text.
    Returns True if row 2 has Arabic content, False otherwise.
    """
    row2_cells = list(ws[2])
    
    # Check if at least one cell in row 2 has Arabic text
    for cell in row2_cells:
        if cell.value and contains_arabic(str(cell.value)):
            return True
    
    return False


def get_guidance_for_column(column_name):
    """
    Get Arabic guidance text for a column based on its name.
    Uses intelligent pattern matching to determine the appropriate guidance.
    """
    if not column_name:
        return "أدخل البيانات المطلوبة"
    
    column_lower = str(column_name).lower()
    
    # Comprehensive guidance mappings based on column patterns
    guidance_patterns = {
        # IDs and Codes
        r'(code|id|رمز|معرف)': "أدخل رمز/معرف فريد (مثال: 001، ABC-123)",
        r'sectioncode|section_code': "أدخل رمز القطاع/المقطع",
        r'intersectioncode': "أدخل رمز التقاطع",
        r'samplecode|sample_code': "أدخل رمز العينة",
        r'regioncode': "أدخل رمز المنطقة",
        r'distresscode|distress_code': "أدخل رمز العيب",
        r'distressno|distress_no': "أدخل رقم نوع العيب",
        r'bridgeid|bridge_id': "أدخل رمز تعريف الجسر",
        
        # Dates and Times
        r'(date|تاريخ)': "أدخل التاريخ بصيغة DD/MM/YYYY",
        r'surveydate': "أدخل تاريخ المسح بصيغة DD/MM/YYYY",
        r'completiondate|completion_date': "أدخل تاريخ الإنجاز المتوقع",
        r'time|وقت': "أدخل الوقت بصيغة HH:MM",
        
        # Location and GPS
        r'(location|موقع)': "أدخل الموقع أو الإحداثيات",
        r'(gps|coordinates|إحداثيات)': "أدخل خط الطول والعرض (Lat, Long)",
        r'(lat|latitude)': "أدخل خط العرض (Latitude)",
        r'(long|longitude)': "أدخل خط الطول (Longitude)",
        
        # Streets and Roads
        r'street|شارع': "أدخل اسم أو رقم الشارع",
        r'substreetnumber': "أدخل رقم الشارع الفرعي",
        r'substreetname': "أدخل اسم الشارع الفرعي",
        r'lane|مسار': "أدخل رقم المسار (L1, L2, إلخ)",
        r'direction|اتجاه': "أدخل اتجاه الحركة (شمال، جنوب، شرق، غرب)",
        
        # Measurements - Dimensions
        r'length|طول': "أدخل الطول بالمتر أو السنتيمتر",
        r'width|عرض': "أدخل العرض بالمتر أو السنتيمتر",
        r'height|ارتفاع': "أدخل الارتفاع بالمتر",
        r'area|مساحة': "أدخل المساحة بالمتر المربع",
        r'substreetarea': "أدخل مساحة الشارع الفرعي بالمتر المربع",
        r'distressarea|distress_area|distess_area': "أدخل مساحة العيب بالمتر المربع",
        r'lane_area': "أدخل مساحة المسار بالمتر المربع",
        r'span.*length': "أدخل طول البحر/الفتحة بالمتر",
        
        # Quantities and Counts
        r'quantity|كمية': "أدخل الكمية بالوحدة المناسبة",
        r'casualties|إصابات': "أدخل عدد الإصابات أو الوفيات",
        
        # Status and Conditions
        r'(status|حالة)': "أدخل الحالة (جيد، متوسط، سيء، إلخ)",
        r'severity': "أدخل درجة الخطورة (Low=منخفض، Medium=متوسط، High=عالي)",
        r'condition': "أدخل حالة الجسر (ممتاز، جيد، متوسط، سيء)",
        r'workingstatus|working_status': "أدخل حالة التشغيل (يعمل، معطل، يحتاج صيانة)",
        
        # Descriptions and Notes
        r'(description|وصف)': "أدخل وصف تفصيلي",
        r'(notes|ملاحظات)': "أدخل أي ملاحظات إضافية",
        
        # Bridge-specific
        r'bridge.*type|نوع.*جسر': "أدخل نوع الجسر (خرساني، معدني، إلخ)",
        r'load.*capacity|حمولة': "أدخل الحمولة القصوى بالطن",
        
        # Traffic Safety
        r'accident.*type': "أدخل نوع الحادث (اصطدام، انقلاب، دهس، إلخ)",
        r'weather|طقس': "أدخل حالة الطقس (صحو، ممطر، ضبابي، إلخ)",
        
        # Lighting
        r'light.*type|نوع.*إنارة': "أدخل نوع الإنارة (LED، صوديوم، هاليد معدني، إلخ)",
        r'power|قدرة': "أدخل القدرة بالواط (W)",
        r'pole.*material': "أدخل مادة العمود (حديد، ألمنيوم، خرسانة)",
        
        # Speed
        r'speed|سرعة': "أدخل السرعة بالكيلومتر/ساعة (km/h)",
        r'vehicle.*type|نوع.*مركبة': "أدخل نوع المركبة (سيارة خاصة، شاحنة، حافلة، دراجة)",
        
        # Improvements
        r'improvement.*type': "أدخل نوع التحسين المطلوب",
        r'cost|تكلفة': "أدخل التكلفة التقديرية بالريال",
        r'priority|أولوية': "أدخل درجة الأولوية (عالية، متوسطة، منخفضة)",
        
        # Technical measurements (FWD, GPR, IRI, SKID)
        r'd\d+': "أدخل قراءة جهاز FWD بالميكرون (مثال: 310.5)",
        r'layer\d+': "أدخل سمك الطبقة بالسنتيمتر",
        r'col\d+': "أدخل البيانات المطلوبة",
        r'iri': "أدخل قيمة IRI (مؤشر الخشونة الدولي) بوحدة m/km",
        r'mu|friction': "أدخل قيمة معامل الاحتكاك/مقاومة الانزلاق",
    }
    
    # Try to match patterns
    for pattern, guidance in guidance_patterns.items():
        if re.search(pattern, column_lower):
            return guidance
    
    # Default guidance
    return "أدخل البيانات المطلوبة"


def add_guidance_row_to_file(filepath):
    """
    Add a guidance row to an Excel file if it doesn't already have one.
    
    Args:
        filepath: Path to the Excel file
        
    Returns:
        str: Status message
    """
    filename = os.path.basename(filepath)
    print(f"\n{'='*60}")
    print(f"Processing: {filename}")
    print(f"{'='*60}")
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Check if row 2 already has Arabic guidance
        if has_arabic_guidance_row(ws):
            print(f"⊘ Skipped: {filename} already has Arabic guidance in row 2")
            wb.close()
            return f"SKIPPED: {filename} (already has guidance)"
        
        # Get headers from row 1
        headers = [cell.value for cell in ws[1]]
        print(f"Headers found: {len([h for h in headers if h])} columns")
        
        # Insert a new row at position 2
        ws.insert_rows(2)
        print(f"Inserted new row at position 2")
        
        # Add guidance text to each column in row 2
        guidance_added = 0
        for col_idx, header in enumerate(headers, start=1):
            if header:
                # Get the appropriate guidance text
                guidance_text = get_guidance_for_column(header)
                
                # Set the guidance text in row 2
                cell = ws.cell(row=2, column=col_idx)
                cell.value = guidance_text
                
                # Format the cell as specified:
                # - Font: size 9, italic, dark gray (#404040)
                # - Fill: light gray (#D3D3D3)
                # - Alignment: center horizontal and vertical, wrap text
                cell.font = Font(size=9, italic=True, color="404040")
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                guidance_added += 1
        
        # Adjust row height for row 2 to accommodate wrapped text
        ws.row_dimensions[2].height = 30
        
        print(f"Added guidance for {guidance_added} columns")
        
        # Save the modified workbook
        wb.save(filepath)
        print(f"✓ Successfully saved {filename}")
        
        wb.close()
        return f"SUCCESS: {filename} ({guidance_added} columns)"
        
    except Exception as e:
        print(f"✗ Error processing {filename}: {str(e)}")
        return f"ERROR: {filename} - {str(e)}"


def main():
    """Main function to process all Excel files in the repository."""
    print("="*60)
    print("Adding Arabic Guidance Rows to Excel Files")
    print("إضافة صفوف الإرشادات العربية لملفات Excel")
    print("="*60)
    
    # Find all .xlsx files in the current directory
    xlsx_files = glob.glob("*.xlsx")
    
    if not xlsx_files:
        print("\n⚠ No .xlsx files found in the current directory")
        return
    
    print(f"\nFound {len(xlsx_files)} Excel file(s)")
    print(f"Files: {', '.join(xlsx_files)}\n")
    
    # Process each file
    results = {
        'success': [],
        'skipped': [],
        'error': []
    }
    
    for filepath in xlsx_files:
        result = add_guidance_row_to_file(filepath)
        
        if result.startswith('SUCCESS'):
            results['success'].append(result)
        elif result.startswith('SKIPPED'):
            results['skipped'].append(result)
        else:
            results['error'].append(result)
    
    # Print summary report
    print("\n" + "="*60)
    print("SUMMARY REPORT / تقرير ملخص")
    print("="*60)
    print(f"\n✓ Successfully processed: {len(results['success'])} file(s)")
    for r in results['success']:
        print(f"  - {r}")
    
    print(f"\n⊘ Skipped (already have guidance): {len(results['skipped'])} file(s)")
    for r in results['skipped']:
        print(f"  - {r}")
    
    if results['error']:
        print(f"\n✗ Errors: {len(results['error'])} file(s)")
        for r in results['error']:
            print(f"  - {r}")
    
    print("\n" + "="*60)
    print("Processing complete! / اكتملت المعالجة!")
    print("="*60)


if __name__ == '__main__':
    main()
