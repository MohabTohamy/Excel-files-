#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script to add Arabic guidance rows to Excel template files.
This script inserts a new row (row 2) with Arabic instructions for each column.

Usage:
    python3 add_guidance_rows.py

Requirements:
    pip install openpyxl
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Define guidance text for each file
guidance_mappings = {
    'قالب-fwd.xlsx': {
        'SectionCode': 'أدخل رمز القطاع/المقطع (مثال: 0120010601014020001)',
        'Lane': 'أدخل رقم المسار (L1, L2, إلخ)',
        'D0': 'أدخل قراءة جهاز FWD بالميكرون (مثال: 310.5)',
    },
    
    'قالب-gpr.xlsx': {
        'SectionCode': 'أدخل رمز القطاع/المقطع',
        'Layer1': 'أدخل سمك الطبقة الأولى بالسنتيمتر',
        # For Col1-Col32, we'll use a default message
        'default': 'أدخل البيانات المطلوبة',
    },
    
    'قالب-iri.xlsx': {
        'SectionCode': 'أدخل رمز القطاع/المقطع',
        'Lane': 'أدخل رقم المسار (L1, L2, إلخ)',
        'IRI': 'أدخل قيمة IRI (مؤشر الخشونة الدولي) بوحدة m/km',
    },
    
    'قالب-skid.xlsx': {
        'SectionCode': 'أدخل رمز القطاع/المقطع',
        'Lane': 'أدخل رقم المسار (L1, L2, إلخ)',
        'Mu': 'أدخل قيمة معامل الاحتكاك/مقاومة الانزلاق',
    },
    
    'قالب-عيوب-التقاطعات.xlsx': {
        'IntersectionCode': 'أدخل رمز التقاطع',
        'SampleCode': 'أدخل رمز العينة',
        'Street1': 'أدخل رقم الشارع الأول',
        'Street2': 'أدخل اسم الشارع الثاني',
        'DistressNo': 'أدخل رقم نوع العيب',
        'Severity': 'أدخل درجة الخطورة (L=منخفض، M=متوسط، H=عالي)',
        'DistessArea': 'أدخل مساحة العيب بالمتر المربع',
        'Length': 'أدخل طول العيب بالسنتيمتر',
        'Width': 'أدخل عرض العيب بالسنتيمتر',
        'SubstreetArea': 'أدخل مساحة الشارع الفرعي',
    },
    
    'قالب-عيوب-الطرق-الرئيسية.xlsx': {
        'Section_Code': 'أدخل رمز القطاع',
        'Lane': 'أدخل رمز المسار (L, R, إلخ)',
        'Distress_No': 'أدخل رقم نوع العيب',
        'Severity': 'أدخل درجة الخطورة (Low, Medium, High)',
        'Lane_Area': 'أدخل مساحة المسار',
        'Length': 'أدخل الطول',
        'Width': 'أدخل العرض',
        'Distress_Area': 'أدخل مساحة العيب',
        'Distress_Length': 'أدخل طول العيب',
        'Distress_Width': 'أدخل عرض العيب',
        'SampleCode': 'أدخل رمز العينة',
        'SurveyDate': 'أدخل تاريخ المسح',
        'default': 'أدخل البيانات المطلوبة',
    },
    
    'قالب-عيوب-الطرق-الفرعية.xlsx': {
        'RegionCode': 'أدخل رمز المنطقة',
        'SubstreetNumber': 'أدخل رقم الشارع الفرعي',
        'SubstreetName': 'أدخل اسم الشارع الفرعي',
        'DistressCode': 'أدخل رمز العيب',
        'Severity': 'أدخل درجة الخطورة (Low, Medium, High)',
        'DistressArea': 'أدخل مساحة العيب بالمتر المربع',
        'Length': 'أدخل طول العيب بالمتر',
        'Width': 'أدخل عرض العيب بالمتر',
        'SubstreetArea': 'أدخل مساحة الشارع الفرعي',
    },
}


def add_guidance_row(filename):
    """
    Add a guidance row to an Excel file.
    
    Args:
        filename: Name of the Excel file to modify
    """
    print(f"\nProcessing: {filename}")
    
    # Load workbook
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Get headers from row 1
    headers = [cell.value for cell in ws[1]]
    print(f"Headers found: {headers}")
    
    # Insert a new row at position 2
    ws.insert_rows(2)
    
    # Get guidance mapping for this file
    guidance = guidance_mappings.get(filename, {})
    default_guidance = guidance.get('default', 'أدخل البيانات')
    
    # Add guidance text to each column in row 2
    for col_idx, header in enumerate(headers, start=1):
        if header:
            # Get the appropriate guidance text
            guidance_text = guidance.get(header, default_guidance)
            
            # Set the guidance text in row 2
            cell = ws.cell(row=2, column=col_idx)
            cell.value = guidance_text
            
            # Format the cell: smaller font, italic, light gray background
            cell.font = Font(size=9, italic=True, color="666666")
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    
    # Adjust row height for row 2 to accommodate wrapped text
    ws.row_dimensions[2].height = 30
    
    # Save the modified workbook
    wb.save(filename)
    print(f"✓ Successfully added guidance row to {filename}")
    
    wb.close()


def main():
    """Main function to process all Excel files."""
    files = [
        'قالب-fwd.xlsx',
        'قالب-gpr.xlsx',
        'قالب-iri.xlsx',
        'قالب-skid.xlsx',
        'قالب-عيوب-التقاطعات.xlsx',
        'قالب-عيوب-الطرق-الرئيسية.xlsx',
        'قالب-عيوب-الطرق-الفرعية.xlsx',
    ]
    
    print("="*60)
    print("Adding Arabic guidance rows to Excel files")
    print("="*60)
    
    for filename in files:
        try:
            add_guidance_row(filename)
        except Exception as e:
            print(f"✗ Error processing {filename}: {str(e)}")
    
    print("\n" + "="*60)
    print("Processing complete!")
    print("="*60)


if __name__ == '__main__':
    main()
